import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook, Workbook

class General_Functions:

    def __init__(self):
        pass

    def create_and_load_workbook(self, workbook_path):
        if os.path.exists(workbook_path)!=True:
            workbook = Workbook()
            workbook.save(workbook_path)
        return load_workbook(workbook_path)
    
    def write_data(self, workbook_path, df):
        experiment_name, workbook = self.set_experiment_name(workbook_path)
        
        writer = pd.ExcelWriter(workbook_path, engine = 'openpyxl')
        writer.book = workbook
        
        df.to_excel(writer, sheet_name=experiment_name)

        writer.save()
        writer.close()

    def set_experiment_name(self, workbook_path):
        workbook = self.create_and_load_workbook(workbook_path)
        sheet_name_splitted_array = workbook.sheetnames[-1].split("_")
        
        experiment_name = ""
        if len(sheet_name_splitted_array) == 1:
            sheet_to_delete = workbook.get_sheet_by_name(sheet_name_splitted_array[0])
            workbook.remove_sheet(sheet_to_delete)
            experiment_name = "Experiment_1"
        else:
            experiment_name = "Experiment_" + str(int(sheet_name_splitted_array[1]) + 1)
        
        return experiment_name, workbook
    
    def get_experiment_name(self, workbook_path):
        workbook = self.create_and_load_workbook(workbook_path)
        return workbook.sheetnames[-1]
       
    def save_excel(self, data_dir_path, df):
        self.create_directory(data_dir_path)
        workbook_path = data_dir_path + "/Data.xlsx"
        self.write_data(workbook_path, df)
    
    def save_plot(self, data_dir_path, plot_type):
        exp_dir_path = data_dir_path + "/Experiments/" + plot_type
        self.create_directory(exp_dir_path)
        
        workbook_path = data_dir_path + "/Data.xlsx"
        experiment_name = self.get_experiment_name(workbook_path)
        
        return exp_dir_path, experiment_name
    
    def find_images_to_plot(self, min_label, max_label, data, similarity, labels):
        images = []
        for i in range(min_label, max_label):
            image_indices_and_max_similarities = []
            
            valid_indices = np.where(labels == i)[0]
            for index in valid_indices:
                cur_max_similarity = similarity[index][np.argmax(similarity[index])]
                image_indices_and_max_similarities.append((index, cur_max_similarity))

            sorted_image_indices_and_max_similarities = sorted(image_indices_and_max_similarities, key=lambda x: x[1], reverse=True)
            desired_similarity = sorted_image_indices_and_max_similarities[0][1]
            
            for index, max_similarity in sorted_image_indices_and_max_similarities:
                if(max_similarity <= desired_similarity):
                    images.append(data[index])
                    desired_similarity = desired_similarity - 0.1 * desired_similarity
                if(len(images) % 10 == 0):
                    break
        
        return images          
        
    def create_directory(self, directory_path):
        if not os.path.exists(directory_path):
            os.makedirs(directory_path)
            print(f"Directory '{directory_path}' created successfully.")
        else:
            print(f"Directory '{directory_path}' already exists.")    
    
